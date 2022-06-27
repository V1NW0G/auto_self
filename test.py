from openpyxl import load_workbook
import pyperclip
import docx 
import os

def fileName():
    file = input("Which file will be use ? ")
    file = file + '.xlsx'
    return file

filename = fileName()
wb = load_workbook(filename)
ws = wb['CMLink英国营销网站Channel ID']

# for value in ws.iter_rows():
#     col = [v.value for v in value]
#     print(f"Select * Where ")
# row = 1
# column = 1
# data1 = ws.cell(row, column).value
# data2 = ws.cell(row, column+1).value
# data3 = ws.cell(row, column+2).value
# print(ws.max_column)
print("The table has updated to " + str(ws.max_row) + "\n")

def countnumber():
    text = input("How many you want to add ? ")
    # print(int(text))
    # print(type(int(text)))
    i = int(text)
    l = []
    l0 = []
    row = ws.max_row - i 
    for i in range(0, i + 1):   
        r = [o.value for o in ws[row]]
        row += 1 
        # r[6] = ''.join(r[6]).replace('新增','').split()
        # new_data = f"{r[3]} {r[4]} {r[5]} {r[6]}"
        tittle = f"--{r[3]}"
        new_data = f"values ('{r[3]}', '{r[3]}', '{r[4]}', '{r[5]}', '1', '1', '1', '10001', 'admin', '0000', '0000', '0000', '英国', null, null, null, null, null, to_date('{r[6]}', 'dd-mm-yyyy'), '74', to_date('{r[6]}', 'dd-mm-yyyy'), '74', null, null, null, null);"
        l0.append(tittle)
        l.append(new_data)

    sourceFile = open('result.txt', 'w')
    l = [s.replace("新增","") for s in l]
    for i in range(0,i):
        print("-- CRM insert", file = sourceFile)
        print(l0[i + 1] , file = sourceFile)
        print("insert into comn.gn_channel (CHNL_ID, CHNL_CODE, CHNL_NAME, CHNL_DESC, STATE, CHNL_KIND_ID, SELF_FLAG, DEPART_ID, MANAGER_STAFF_ID, PROVINCE_CODE, CITY_CODE, AREA_CODE, ADDRESS, POSTCODE, LONGITUDE, LATITUDE, CONTACT_TEL, SUPER_CHNL_ID, CREATE_TIME, CREATE_OPER_ID, UPDATE_TIME, UPDATE_OPER_ID, REMARK, MVNO_CHNL_ID, FEE_CLOSE_FLAG, AGENT_ID)", file = sourceFile)       
        print(l[i + 1] + "\n", file = sourceFile)
    sourceFile.close()


    sourceFile = open('result2.txt', 'w')
    l = [s.replace("新增","") for s in l]
    for i in range(0,i + 1):
        print(l0[i + 1] , file = sourceFile)
        print("-- 账管 insert", file = sourceFile)
        print("-- =1st table=", file = sourceFile)
        print("-- =1st table=", file = sourceFile)
        print("-- -> UK", file = sourceFile)
        print("insert into comn.gn_channel (CHNL_ID, CHNL_CODE, CHNL_NAME, CHNL_DESC, STATE, CHNL_KIND_ID, SELF_FLAG, DEPART_ID, MANAGER_STAFF_ID, PROVINCE_CODE, CITY_CODE, AREA_CODE, ADDRESS, POSTCODE, LONGITUDE, LATITUDE, CONTACT_TEL, SUPER_CHNL_ID, CREATE_TIME, CREATE_OPER_ID, UPDATE_TIME, UPDATE_OPER_ID, REMARK, MVNO_CHNL_ID, FEE_CLOSE_FLAG, AGENT_ID)insert into  param.bs_merchant_detail(channel_id,para_type,plat_type,business_type,pay_merchant_id,merchant_key,para_value1,para_value2,para_value3,para_value4,state,create_date,remarks)", file = sourceFile)       
        print(l[i + 1] + "\n", file = sourceFile)
        print("-- =2nd table= ", file = sourceFile)
        print("-- UK", file = sourceFile)
    sourceFile.close()
    
    
    
# def getResult():
#     # print("this is about " + data1 +  " blabalbabfdafdjgfhsaf " + data2 + data3)
#     sourceFile = open('result.txt', 'w')
#     print("this is about " + data1 +  " blabalbabfdafdjgfhsaf " + data2 + data3, file = sourceFile)
#     sourceFile.close()

def chooseFile():
    choose = input("Which File u want ? Enter 0 to exit ")
    def function1():
        os.startfile('result.txt')
        
    def function2():
        os.startfile('result2.txt')

    def function3():
        print("3")

    if int(choose) == 1:
        function1()
        chooseFile()

    if int(choose) == 2:
        function2()
        chooseFile()
    
    if int(choose) == 0:
        exit()


countnumber()
chooseFile()

# file = fileName()
# print(file)
# pyperclip.copy()
